"""
Generate Laporan Akhir Desk Review (Excel Rekap).
Sheet 1: Ringkasan Eksekutif per RS
Sheet 2: Master Data Rincian per SEP
Sheet 3: Rekap Temuan KNAVP per RS (baru)

Menggunakan kolom KNAVP v2: knavp_skor, tingkat_risiko, keputusan_sistem,
jumlah_beda_dual_coding, ccl_label.
"""
import os
import json
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

from modules.db_manager import get_recap_desk_review, save_generated_report

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'exports', 'rekap')

# ── Palette ──────────────────────────────────────────────────────────────────
BLUE_DARK  = '1E3A5F'
BLUE_MED   = '0369A1'
GREEN_DARK = '166534'
RED_DARK   = '991B1B'
AMBER_DARK = '92400E'
WHITE      = 'FFFFFF'
LIGHT_BLUE = 'DBEAFE'
LIGHT_RED  = 'FEF2F2'
LIGHT_AMB  = 'FFFBEB'

thin  = Side(style='thin')
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _hdr_style(fill_color=BLUE_DARK):
    return {
        'fill': PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
        'font': Font(color=WHITE, bold=True, size=9),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': border,
    }


def _apply(cell, **props):
    for k, v in props.items():
        setattr(cell, k, v)


def _set_row(ws, row_num, row_data, fill=None, bold=False, num_fmt=None):
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.font = Font(bold=bold, size=8)
        if fill:
            cell.fill = fill
        if num_fmt and isinstance(val, (int, float)):
            cell.number_format = num_fmt


def _auto_width(ws, max_width=55):
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)


def generate_recap_excel():
    print("=" * 60)
    print("Generating Laporan Akhir Desk Review (Excel)...")
    print("=" * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    dr_data = get_recap_desk_review()
    print(f"Ditemukan {len(dr_data)} record Desk Review.")

    if not dr_data:
        print("Data kosong.")
        return None

    # ── Parse semua baris ────────────────────────────────────────────────────
    master_data = []
    for row in dr_data:
        try:
            form_data = json.loads(row.get('tindakan_reviewer') or '{}')
        except Exception:
            form_data = {}

        tarif_ina = float(row.get('tarif_inacbg') or 0)
        tarif_rs  = float(row.get('tarif_rs') or 0)
        selisih   = tarif_ina - tarif_rs

        # KNAVP v2 fields (dari kolom DB atau form_data)
        knavp_skor    = form_data.get('knavp_skor', row.get('knavp_skor', 0)) or 0
        tingkat_risiko = form_data.get('tingkat_risiko', row.get('tingkat_risiko', '-')) or '-'
        keputusan_sistem = (form_data.get('keputusan_sistem')
                            or form_data.get('keputusan')
                            or row.get('keputusan_sistem')
                            or form_data.get('keputusan_reviewer', 'Belum Direview'))
        jumlah_beda_dc = form_data.get('jumlah_beda_dual_coding', row.get('jumlah_beda_dual_coding', 0)) or 0
        ccl_label      = form_data.get('ccl_label', row.get('ccl_label', '-')) or '-'

        master_data.append({
            'Nomor SEP':                 row.get('sep'),
            'Kode RS':                   row.get('kode_rs'),
            'Nama RS':                   row.get('nama_rs'),
            'Kelas RS':                  row.get('kelas'),
            'Regional':                  row.get('regional'),
            'Kode INA-CBG':              row.get('inacbg'),
            'Deskripsi INA-CBG':         row.get('deskripsi_inacbg'),
            'Kode iDRG':                 row.get('idrg_code', '-'),
            'CCL (iDRG)':                ccl_label,
            'Tarif INA-CBG (Rp)':        tarif_ina,
            'Tarif RS Standar (Rp)':     tarif_rs,
            'Selisih Tarif (Rp)':        selisih,
            'Skor KNAVP':                knavp_skor,
            'Tingkat Risiko':            tingkat_risiko,
            'Perbedaan Dual Coding':     jumlah_beda_dc,
            'Keputusan Sistem':          keputusan_sistem,
            'Reviewer':                  row.get('reviewer_name') or form_data.get('reviewer_name', 'System'),
            'Tanggal Review':            str(row.get('updated_at', ''))[:10] or form_data.get('tanggal_review', '-'),
            'Analisis / Alasan':         form_data.get('alasan_keputusan', form_data.get('analisis_reviewer', '-')),
        })

    df = pd.DataFrame(master_data)

    # ── Sheet 1: Ringkasan per RS ────────────────────────────────────────────
    summary_rows = []
    for rs_code, g in df.groupby('Kode RS'):
        total = len(g)
        onsite   = g['Keputusan Sistem'].str.contains('On-Site', na=False).sum()
        sampling = g['Keputusan Sistem'].str.contains('Sampling', na=False).sum()
        monitor  = total - onsite - sampling
        beda_dc  = g['Perbedaan Dual Coding'].sum()
        avg_skor = round(g['Skor KNAVP'].mean(), 1)
        max_skor = g['Skor KNAVP'].max()
        sel_total = g['Selisih Tarif (Rp)'].sum()
        summary_rows.append({
            'Kode RS':                   rs_code,
            'Nama RS':                   g['Nama RS'].iloc[0],
            'Total Kasus Audit':         total,
            'Rekomendasi On-Site Audit': onsite,
            'Rekomendasi Sampling':      sampling,
            'Monitoring':                monitor,
            'Rata-rata Skor KNAVP':      avg_skor,
            'Skor KNAVP Tertinggi':      max_skor,
            'Total Perbedaan Dual Coding': int(beda_dc),
            'Total Potensi Selisih Tarif (Rp)': sel_total,
        })
    df_summary = pd.DataFrame(summary_rows)

    # ── Sheet 3: Rekap Temuan per Kategori Risiko ────────────────────────────
    risk_rows = []
    for tk, g2 in df.groupby('Tingkat Risiko'):
        risk_rows.append({
            'Tingkat Risiko':    tk,
            'Jumlah Kasus':      len(g2),
            '% dari Total':      f"{len(g2)/max(len(df),1)*100:.1f}%",
            'Rata-rata Skor':    round(g2['Skor KNAVP'].mean(), 1),
            'Avg Perbedaan DC':  round(g2['Perbedaan Dual Coding'].mean(), 1),
        })
    df_risk = pd.DataFrame(risk_rows) if risk_rows else pd.DataFrame(
        columns=['Tingkat Risiko','Jumlah Kasus','% dari Total','Rata-rata Skor','Avg Perbedaan DC']
    )

    # ── Simpan ke Excel ──────────────────────────────────────────────────────
    ts       = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Laporan_Akhir_Desk_Review_{ts}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)

    print(f"Menyimpan: {filename} ...")

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Ringkasan Eksekutif', index=False)
        df.to_excel(writer, sheet_name='Master Data (Rincian)', index=False)
        df_risk.to_excel(writer, sheet_name='Rekap Risiko KNAVP', index=False)

        wb = writer.book

        # Styling per sheet
        sheet_styles = {
            'Ringkasan Eksekutif': BLUE_DARK,
            'Master Data (Rincian)': BLUE_MED,
            'Rekap Risiko KNAVP': '7C3AED',
        }

        for sheet_name, hdr_color in sheet_styles.items():
            ws = wb[sheet_name]
            hdr = _hdr_style(hdr_color)
            for cell in ws[1]:
                _apply(cell, **hdr)

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border
                    cell.font   = Font(size=8)
                    # Colour-code risiko rows
                    if sheet_name == 'Master Data (Rincian)':
                        tk_val = str(ws.cell(row=cell.row, column=14).value or '')
                        if tk_val == 'Tinggi':
                            cell.fill = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type='solid')
                        elif tk_val == 'Sedang':
                            cell.fill = PatternFill(start_color=LIGHT_AMB, end_color=LIGHT_AMB, fill_type='solid')
                    # Currency format
                    if isinstance(cell.value, (int, float)) and abs(cell.value) > 1000:
                        cell.number_format = '#,##0'

            ws.freeze_panes = 'A2'
            _auto_width(ws)

    print(f"SELESAI! Laporan tersimpan: {filepath}")

    rel_path = f"exports/rekap/{filename}"
    save_generated_report('REKAP_EXCEL', filename, rel_path)

    return filepath


if __name__ == '__main__':
    generate_recap_excel()
